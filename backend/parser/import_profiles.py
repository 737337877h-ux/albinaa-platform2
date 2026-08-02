# -*- coding: utf-8 -*-
"""
Import profile detection & parsers — APMS v1.2.0 (PR 2)
========================================================
يدعم خمسة ملفات تعريف استيراد:
  1. CUSTOMER_MASTER              — بيانات العملاء الأساسية
  2. CUSTOMER_BALANCE_SUMMARY     — ملخص أرصدة العملاء (حسب العملة)
  3. DEBT_AGING_SUMMARY           — تقسيم الأعمار المجمّع (حسب العملة)
  4. DEBT_AGING_DETAILS           — تقسيم الأعمار التفصيلي (عميل-بعميل)
  5. CUSTOMER_STATEMENT_DETAILS   — كشف الحساب التحليلي (المُختبر سابقًا
                                   في albinaa_parser.py — يُحوَّل إليه)

الصيغ المدعومة (قرار R1 المعتمد):
  - XLSX / XLSM عبر openpyxl (read_only + data_only).
  - XLS نصي مفصول بـ Tab بترميز CP1256 (Windows-1256) افتراضيًا، مع دعم
    BOM لـ UTF-8 / UTF-16. فك الترميز صارم (errors='strict'): أي فشل يُرفض
    بخطأ واضح ولا تُفسَد البيانات العربية أبدًا.
  - XLS الثنائية القديمة (OLE2) غير مدعومة — خطأ واضح يوجّه للتصدير كـ xlsx.

التعرّف على نوع الملف قائم على الترويسة (Heuristic):
  - صف ترؤيسة فيه ≥2 تسمية معروفة → ملف جدولي (ماستر/أرصدة/أعمار).
  - أعمدة تقسيم الأعمار (0-30, أكثر من 120, ...) → أعمار (تفصيلي إن وُجد
    كود العميل، مجمّع وإلا).
  - العملة + الرصيد + كود العميل → ملخص أرصدة.
  - كود العميل + اسم العميل → بيانات العملاء.
  - لا ترويسة جدولية + علامات الكتل (رقم العميل / التاريخ) → كشف الحساب.
"""
import re
import unicodedata

from openpyxl import load_workbook

PROFILE_STATEMENT = 'CUSTOMER_STATEMENT_DETAILS'
PROFILE_MASTER = 'CUSTOMER_MASTER'
PROFILE_BALANCE = 'CUSTOMER_BALANCE_SUMMARY'
PROFILE_AGING_SUMMARY = 'DEBT_AGING_SUMMARY'
PROFILE_AGING_DETAILS = 'DEBT_AGING_DETAILS'

CURRENCY_MAP = {'YR': 'YER', 'SR': 'SAR', '$': 'USD'}

OLE2_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'


# ----------------------------------------------------------------------------
# تطبيع / تحويلات
# ----------------------------------------------------------------------------
def normalize_text(s):
    """توحيد نص للمقارنة: NFKC + طي المسافات + Trim."""
    if s is None:
        return ''
    s = unicodedata.normalize('NFKC', str(s))
    return re.sub(r'\s+', ' ', s).strip()


DIGIT_TRANS = str.maketrans('٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹', '01234567890123456789')


def normalize_digits(s):
    """تحويل الأرقام العربية/الفارسية إلى ASCII (الترويسة والقيم)."""
    if s is None:
        return ''
    return str(s).translate(DIGIT_TRANS)


def to_number(v):
    """تحويل خلية إلى رقم عائم. None عند عدم الصلاحية. يدعم الفواصل العربية."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    s = normalize_digits(str(v)).strip()
    s = s.replace('٫', '.').replace('٬', '')   # فاصلة عشرية/آلاف عربية
    s = re.sub(r'[\s\u00a0,]', '', s)
    if s in ('', '-', '--', '.', '-'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ----------------------------------------------------------------------------
# القراءة (XLSX/XLSM عبر openpyxl + XLS نصي بترميز CP1256)
# ----------------------------------------------------------------------------
def decode_xls_text(path):
    """فك ترميز ملف XLS نصي: BOM → UTF-8/UTF-16 وإلا CP1256 صارم (R1)."""
    with open(path, 'rb') as f:
        data = f.read()
    try:
        if data[:3] == b'\xef\xbb\xbf':
            text = data[3:].decode('utf-8')
        elif data[:2] == b'\xff\xfe':
            text = data[2:].decode('utf-16-le')
        elif data[:2] == b'\xfe\xff':
            text = data[2:].decode('utf-16-be')
        else:
            text = data.decode('cp1256')
    except UnicodeDecodeError as e:
        raise ValueError(
            'تعذر فك ترميز الملف — الترميز الافتراضي CP1256 (Windows-1256). '
            'تأكد أن الملف نصي مفصول بـ Tab ولا يحتوي محارف غير قابلة للفك. '
            f'(تفصيل: {e})')
    # حماية R1: لا إفساد صامت — رفض الملفات التي تحوي محارف تحكم (نصوص ثنائية/
    # مقطوعة) بدل إنتاج نص عربي مشوّه.
    controls = sum(1 for ch in text if ord(ch) < 32 and ch not in '\t\r\n')
    if '\x00' in text or controls > 1:
        raise ValueError(
            'الملف لا يبدو نصًا عربيًا سليمًا بـ CP1256 — يحتوي محارف تحكم. '
            'تأكد أن الملف نصي مفصول بـ Tab (ترميز Windows-1256) أو XLSX سليم')
    return text


def read_table(path):
    """قراءة الملف إلى صفوف قوائم. يعيد (rows, format) حيث format: xlsx|tsv."""
    with open(path, 'rb') as f:
        head = f.read(8)
    ext = path.lower().rsplit('.', 1)[-1] if '.' in path else ''
    if ext in ('xlsx', 'xlsm'):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # read_only يُبقي مقبض الملف مفتوحًا — نغلقه بعد التحميل الكامل.
        wb.close()
        return rows, 'xlsx'
    if ext == 'xls':
        if head[:8] == OLE2_MAGIC:
            raise ValueError(
                'ملف Excel قديم بصيغة ثنائية (OLE2) غير مدعوم — صدّره من النظام '
                'المحاسبي كـ xlsx أو كملف نصي مفصول بـ Tab بترميز Windows-1256')
        text = decode_xls_text(path)
        rows = []
        for ln in text.splitlines():
            ln = ln.rstrip('\r')
            rows.append(ln.split('\t'))
        return rows, 'tsv'
    raise ValueError(
        f'صيغة الملف غير مدعومة: .{ext} — المدعوم: xlsx / xlsm / xls (نصي بـ Tab)')


# ----------------------------------------------------------------------------
# تسميات الأعمدة + كشف أعمدة تقسيم الأعمار
# ----------------------------------------------------------------------------
COL_ALIASES = {
    'customer_code': ['رقم العميل', 'كود العميل', 'كود العميل/المورد', 'الكود',
                      'رقم الكود', 'رقم الحساب', 'حساب العميل', 'كود الحساب'],
    'customer_name': ['اسم العميل', 'الاسم', 'الاسم بالعربي', 'اسم العميل/المورد',
                      'الاسم التجاري', 'اسم المورد'],
    'currency': ['العملة', 'عملة', 'رمز العملة', 'كود العملة'],
    'account_number': ['رقم الحساب', 'رقم الحساب البنكي', 'الحساب البنكي', 'حساب'],
    'phone': ['الهاتف', 'رقم الهاتف', 'الجوال', 'رقم الجوال', 'الهاتف / الجوال',
              'الموبايل', 'رقم الموبايل'],
    'whatsapp': ['واتساب', 'رقم واتساب', 'الواتس'],
    'address': ['العنوان', 'الموقع'],
    'region': ['المنطقة', 'المدينة', 'الحي', 'المحافظة'],
    'customer_type': ['نوع العميل', 'التصنيف', 'قطاع العميل', 'القطاع'],
    'balance': ['الرصيد', 'الرصيد الحالي', 'الرصيد المتبقي', 'المديونية',
                'إجمالي الرصيد', 'الرصيد المدين', 'المبلغ المتبقي', 'صافي الرصيد'],
    'opening': ['الرصيد الافتتاحي', 'الرصيد الإفتتاحي', 'الرصيد السابق'],
    'total': ['الإجمالي', 'المجموع', 'إجمالي الرصيد', 'الإجمالي العام'],
}

# ترتيب الأعمدة الجدولية المستخدم في التعرّف على الترويسة
FLAT_FIELD_ORDER = ['customer_code', 'customer_name', 'currency', 'balance',
                    'total', 'phone', 'address', 'region', 'customer_type']

# مطابقة بالاحتواء (بعد الفشل بالمطابقة التامة) — تسميات قوية لا تُلتبس
SUBSTR_RULES = [
    ('customer_code', ['رقم العميل', 'كود العميل', 'كود الحساب', 'رقم الحساب']),
    ('customer_name', ['اسم العميل', 'اسم المورد']),
    ('currency', ['العملة', 'عملة']),
    ('balance', ['الرصيد', 'المديونية']),
    ('total', ['الإجمالي', 'المجموع']),
    ('phone', ['الجوال', 'الهاتف', 'الموبايل']),
]

BUCKET_SUFFIX = re.compile(r'(يوم|يومًا|يوما|أيام|ايام|شهر|شهرًا|شهرا)\s*$')

# بنى التصدير الطباعية من APMS: تسميات متكررة كبادئة لكل صف + قيم بترتيب ثابت
MASTER_REPEATED_LABELS = [
    'اسم العميل', 'رقم الحساب', 'المجموعة', 'المدينه',
    'رقم التلفون', 'توقيف', 'تاريخ التعامل', 'رقم العميل',
]
MASTER_VALUE_COLS = {
    'customer_code': 8, 'customer_name': 9, 'account_number': 10,
    'customer_group': 11, 'region': 12, 'phone': 13,
    'blocked': 14, 'last_deal_date': 15,
}
AGING_SUMMARY_REPEATED_LABELS = [
    'إجمالي المبلغ المستحق', 'رقم العميل', 'اسم العميل', 'العملة',
    'المبلغ', '0 - 30', '31 - 60', '61 - 90', '91 - 120', '> 120',
]
AGING_SUMMARY_VALUE_COLS = {
    'customer_code': 10, 'customer_name': 11, 'currency': 12,
    'total': 13, 'bucket0': 14,
}
# كتلة الأعمار التفصيلية: القيم تتراجع في عمود فئتها العمرية (4..8)
DETAILS_BUCKET_KEYS = ['0-30', '31-60', '61-90', '91-120', '120+']
DETAILS_BUCKET_COLS = [4, 5, 6, 7, 8]


def bucket_key(label):
    """مفتاح موحد لعمود تقسيم الأعمار (مثل 0-30 أو 120+) أو None."""
    s = normalize_digits(normalize_text(label))
    s = BUCKET_SUFFIX.sub('', s).strip()
    if not s:
        return None
    m = re.match(r'^(من\s+)?(\d+)\s*(إلى|الي|الى|حتى|لحد)\s*(\d+)$', s)
    if m:
        return f'{int(m.group(2))}-{int(m.group(4))}'
    m = re.match(r'^(\d+)\s*[-–—]\s*(\d+)$', s)
    if m:
        return f'{int(m.group(1))}-{int(m.group(2))}'
    m = re.match(r'^(أكثر من|اكثر من|أكبر من|اكبر من)\s*(\d+)$', s)
    if m:
        return f'{int(m.group(2))}+'
    m = re.match(r'^[<>≥]\s*(\d+)$', s)
    if m:
        return f'{int(m.group(1))}+'
    m = re.match(r'^(\d+)\s*\+$', s)
    if m:
        return f'{int(m.group(1))}+'
    # ملاحظة: لا نطابق رقمًا عاريًا (مثل 120) كعمود أعمار — كود عميل مثل 1001
    # يُصنَّف خطأً كعمود تقسيم ويُفسد كشف الملف. الصيغ الأوضح كافية للملفات الحقيقية.
    return None


def classify_header(cell):
    """تصنيف خلية ترويسة → ('field', token) أو ('bucket', key) أو None."""
    s = normalize_text(cell)
    if not s:
        return None
    bk = bucket_key(s)
    if bk:
        return ('bucket', bk)
    for field in FLAT_FIELD_ORDER:
        for a in COL_ALIASES[field]:
            if normalize_text(a) == s:
                return (field, s)
    for field, tokens in SUBSTR_RULES:
        for t in tokens:
            if t in s:
                return (field, s)
    return None


def _is_empty_row(row):
    return not any(c is not None and str(c).strip() != '' for c in row)


def find_header_row(rows, limit=30):
    """صف الترويسة الجدولية (≥2 تسمية معروفة) في أول limit صفًا."""
    for idx, row in enumerate(rows[:limit]):
        if _is_empty_row(row):
            continue
        hits = sum(1 for c in row if classify_header(c))
        if hits >= 2:
            return idx
    return None


def sort_bucket_columns(cols):
    """ترتيب أعمدة الأعمار تصاعديًا حسب الحد الأدنى."""
    def lo(key):
        if '-' in key:
            return int(key.split('-')[0])
        return int(key.rstrip('+'))
    return sorted(cols, key=lambda x: (lo(x[1]), x[1]))


def build_column_map(header):
    """خريطة (field -> col) + قائمة أعمدة الأعمار (col, key)."""
    cmap = {}
    bucket_cols = []
    for i, cell in enumerate(header):
        k = classify_header(cell)
        if k is None:
            continue
        if k[0] == 'bucket':
            bucket_cols.append((i, k[1]))
        elif k[0] != 'account_number':
            cmap.setdefault(k[0], i)
    # 'رقم الحساب' قد يُصنَّف كودًا — إن وُجد عمود آخر واضحًا فهو رقم الحساب
    for i, cell in enumerate(header):
        s = normalize_text(cell)
        if 'رقم الحساب' in s and cmap.get('customer_code') != i:
            cmap['account_number'] = i
            break
    return cmap, bucket_cols


def _label_row(row, labels):
    """الصف يبدأ بتسميات ثابتة (بنيَة التصدير الطباعي من APMS)."""
    if len(row) < len(labels) + 1:
        return False
    for i, lbl in enumerate(labels):
        if normalize_text(row[i]) != normalize_text(lbl):
            return False
    return True


def _is_aging_block_header(row):
    """رأس كتلة تفصيلية: (العملة : $، الرصيد :، رقم العميل : CODE NAME)."""
    if not row:
        return False
    s0 = normalize_text(row[0])
    s1 = normalize_text(row[1]) if len(row) > 1 else ''
    s3 = normalize_text(row[3]) if len(row) > 3 else ''
    return 'العملة' in s0 and 'الرصيد' in s1 and 'رقم العميل' in s3


def _bucket_num(v):
    """رقم عمود أعمار: خلية فارغة → 0.0، نص غير رقمي → None (يُحسب خطأ)."""
    if v is None or str(v).strip() == '':
        return 0.0
    return to_number(v)


# ----------------------------------------------------------------------------
# كشف نوع الملف
# ----------------------------------------------------------------------------
# عناوين فتح كتلة كشف الحساب التحليلي — تنويعان حقيقيان للتصدير:
# 'رقم العميل' (موجَّه للعميل) و'رقم الحساب' (موجَّه لدليل الحسابات، والهوية
# فيه على سطر 'الحساب التحليلي' التالي).
STATEMENT_BLOCK_LABELS = ('رقم العميل', 'رقم الحساب')


def detect_profile(rows):
    # بنى التصدير الطباعي (تُفحص أولًا لأنها تزيّف الكشف الجدولي)
    for row in rows[:8]:
        if _is_aging_block_header(row):
            return PROFILE_AGING_DETAILS
        if _label_row(row, AGING_SUMMARY_REPEATED_LABELS):
            return PROFILE_AGING_SUMMARY
        if _label_row(row, MASTER_REPEATED_LABELS):
            return PROFILE_MASTER
    header_idx = find_header_row(rows)
    if header_idx is not None:
        cmap, bucket_cols = build_column_map(rows[header_idx])
        if bucket_cols:
            return (PROFILE_AGING_DETAILS if 'customer_code' in cmap
                    else PROFILE_AGING_SUMMARY)
        if 'customer_code' in cmap and 'currency' in cmap and 'balance' in cmap:
            return PROFILE_BALANCE
        if 'customer_code' in cmap and 'customer_name' in cmap:
            return PROFILE_MASTER
        raise ValueError(
            'لم يُتعرَّف على بنية الملف — تأكد من ترويسة الأعمدة: '
            'كود العميل + اسم العميل، أو + العملة/الرصيد، أو أعمدة تقسيم الأعمار')
    seen_date = any(row and normalize_text(row[0]) == 'التاريخ' for row in rows)
    seen_block = any(
        row and normalize_text(row[0]) in STATEMENT_BLOCK_LABELS for row in rows)
    if seen_date and seen_block:
        return PROFILE_STATEMENT
    raise ValueError(
        'لم يُتعرَّف على نوع الملف — المدعوم: كشف حساب تحليلي (بنية الكتل)، '
        'بيانات عملاء، ملخص أرصدة، أو تقسيم أعمار')


# ----------------------------------------------------------------------------
# محللات الملفات الجدولية
# ----------------------------------------------------------------------------
def _cell(row, j):
    if j is None or j >= len(row) or row[j] is None:
        return ''
    return str(row[j]).strip()


def _profile_result(profile, rows, header_idx, errors, records, skipped, total_label):
    return {
        'profile': profile,
        'stats': {
            'rows': len(rows),
            'validRows': len(records),
            'errors': len(errors),
            'emptyRowsSkipped': skipped,
        },
        'errors': [
            {'rowNumber': r, 'message': m, 'raw': list(rw)} for r, m, rw in errors
        ],
        'skippedEmptyRows': skipped,
        total_label: records,
    }


def parse_customer_master(rows):
    header_idx = find_header_row(rows)
    if header_idx is None:
        raise ValueError('ملف بيانات العملاء بدون ترويسة أعمدة — لا يمكن تحديد الأعمدة')
    cmap, _ = build_column_map(rows[header_idx])
    if 'customer_code' not in cmap or 'customer_name' not in cmap:
        raise ValueError('ملف بيانات العملاء يفتقد عمودي (رقم العميل / اسم العميل)')
    records = []
    errors = []
    skipped = 0
    seen = set()
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if _is_empty_row(row):
            skipped += 1
            continue
        code = normalize_digits(_cell(row, cmap.get('customer_code')))
        name = _cell(row, cmap.get('customer_name'))
        rownum = i + 1
        if not code:
            errors.append((rownum, 'كود عميل ناقص — الصف مستبعد', row))
            continue
        if code in seen:
            errors.append((rownum, f'كود مكرر داخل الملف ({code}) — الصف مستبعد', row))
            continue
        if not name:
            errors.append((rownum, 'اسم عميل ناقص — الصف مستبعد', row))
            continue
        seen.add(code)
        records.append({
            'rowNumber': rownum,
            'customerCode': code,
            'customerName': name,
            'accountNumber': normalize_digits(_cell(row, cmap.get('account_number'))) or None,
            'phone': _cell(row, cmap.get('phone')) or None,
            'whatsapp': _cell(row, cmap.get('whatsapp')) or None,
            'region': _cell(row, cmap.get('region')) or None,
            'address': _cell(row, cmap.get('address')) or None,
            'customerType': _cell(row, cmap.get('customer_type')) or None,
        })
    return _profile_result(PROFILE_MASTER, rows, header_idx, errors, records, skipped,
                           'customers')


def parse_balance_summary(rows):
    header_idx = find_header_row(rows)
    if header_idx is None:
        raise ValueError('ملف ملخص الأرصدة بدون ترويسة أعمدة — لا يمكن تحديد الأعمدة')
    cmap, _ = build_column_map(rows[header_idx])
    if 'customer_code' not in cmap or 'currency' not in cmap or 'balance' not in cmap:
        raise ValueError('ملف ملخص الأرصدة يفتقد أعمدة (رقم العميل / العملة / الرصيد)')
    records = []
    errors = []
    skipped = 0
    seen = set()
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if _is_empty_row(row):
            skipped += 1
            continue
        code = normalize_digits(_cell(row, cmap.get('customer_code')))
        name = _cell(row, cmap.get('customer_name'))
        ccy_raw = _cell(row, cmap.get('currency'))
        bal = to_number(_cell(row, cmap.get('balance')))
        opening = to_number(_cell(row, cmap.get('opening')))
        rownum = i + 1
        if not code:
            errors.append((rownum, 'كود عميل ناقص — الصف مستبعد', row))
            continue
        if not ccy_raw:
            errors.append((rownum, 'عملة ناقصة — الصف مستبعد', row))
            continue
        if bal is None:
            errors.append((rownum, 'رصيد غير رقمي — الصف مستبعد', row))
            continue
        key = (code, ccy_raw)
        if key in seen:
            errors.append((rownum, f'سطر مكرر ({code} / {ccy_raw}) — الصف مستبعد', row))
            continue
        seen.add(key)
        records.append({
            'rowNumber': rownum,
            'customerCode': code,
            'customerName': name or code,
            'currencyRaw': ccy_raw,
            'currency': CURRENCY_MAP.get(ccy_raw, ccy_raw),
            'balance': bal,
            'openingBalance': opening,
        })
    return _profile_result(PROFILE_BALANCE, rows, header_idx, errors, records, skipped,
                           'balances')


def parse_debt_aging(rows, details):
    header_idx = find_header_row(rows)
    if header_idx is None:
        raise ValueError('ملف تقسيم الأعمار بدون ترويسة أعمدة — لا يمكن تحديد الأعمدة')
    cmap, bucket_cols = build_column_map(rows[header_idx])
    bucket_cols = sort_bucket_columns(bucket_cols)
    if not bucket_cols:
        raise ValueError('ملف تقسيم الأعمار بدون أعمدة تقسيم (مثل 0-30, 31-60, أكثر من 120)')
    if details and 'customer_code' not in cmap:
        raise ValueError('ملف الأعمار التفصيلي يفتقد عمود كود العميل')
    if not details and 'currency' not in cmap:
        raise ValueError('ملف الأعمار المجمّع يفتقد عمود العملة')
    records = []
    errors = []
    skipped = 0
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if _is_empty_row(row):
            skipped += 1
            continue
        code = normalize_digits(_cell(row, cmap.get('customer_code')))
        name = _cell(row, cmap.get('customer_name'))
        ccy_raw = _cell(row, cmap.get('currency'))
        rownum = i + 1
        if not code and details:
            errors.append((rownum, 'كود عميل ناقص — الصف مستبعد', row))
            continue
        if not ccy_raw:
            errors.append((rownum, 'عملة ناقصة — الصف مستبعد', row))
            continue
        buckets = {}
        for j, bk in bucket_cols:
            num = to_number(_cell(row, j))
            if num is None:
                errors.append(
                    (rownum, f'قيمة غير رقمية في عمود الأعمار ({bk}) — تُعتبر صفرًا', row))
                buckets[bk] = 0.0
            else:
                buckets[bk] = num
        total = to_number(_cell(row, cmap.get('total')))
        rec = {
            'rowNumber': rownum,
            'currencyRaw': ccy_raw,
            'currency': CURRENCY_MAP.get(ccy_raw, ccy_raw),
            'buckets': buckets,
            'total': total,
        }
        if details:
            rec['customerCode'] = code
            rec['customerName'] = name or code
        records.append(rec)
    label = 'agingDetails' if details else 'agingSummary'
    profile = PROFILE_AGING_DETAILS if details else PROFILE_AGING_SUMMARY
    return _profile_result(profile, rows, header_idx, errors, records, skipped, label)


def parse_customer_master_repeated(rows):
    """بيانات العملاء بتصدير الطباعة: تسميات مكررة + قيم بترتيب ثابت."""
    records = []
    errors = []
    skipped = 0
    seen = set()
    for i, row in enumerate(rows):
        if _is_empty_row(row):
            skipped += 1
            continue
        if not _label_row(row, MASTER_REPEATED_LABELS):
            errors.append(
                (i + 1, 'سطر خارج بنية بيانات العملاء (التسميات المتكررة غير مكتملة) — الصف مستبعد', row))
            continue
        code = normalize_digits(_cell(row, MASTER_VALUE_COLS['customer_code']))
        name = _cell(row, MASTER_VALUE_COLS['customer_name'])
        rownum = i + 1
        if not code:
            errors.append((rownum, 'كود عميل ناقص — الصف مستبعد', row))
            continue
        if code in seen:
            errors.append((rownum, f'كود مكرر داخل الملف ({code}) — الصف مستبعد', row))
            continue
        if not name:
            errors.append((rownum, 'اسم عميل ناقص — الصف مستبعد', row))
            continue
        seen.add(code)
        records.append({
            'rowNumber': rownum,
            'customerCode': code,
            'customerName': name,
            'accountNumber': normalize_digits(_cell(row, MASTER_VALUE_COLS['account_number'])) or None,
            'customerGroup': _cell(row, MASTER_VALUE_COLS['customer_group']) or None,
            'region': _cell(row, MASTER_VALUE_COLS['region']) or None,
            'phone': _cell(row, MASTER_VALUE_COLS['phone']) or None,
            'blocked': _cell(row, MASTER_VALUE_COLS['blocked']) or None,
            'lastDealDate': _cell(row, MASTER_VALUE_COLS['last_deal_date']) or None,
        })
    return _profile_result(PROFILE_MASTER, rows, 0, errors, records, skipped, 'customers')


def parse_aging_summary_repeated(rows):
    """ملخص تقسيم الأعمار بتصدير الطباعة: تسميات مكررة + قيم ثابتة."""
    records = []
    errors = []
    skipped = 0
    for i, row in enumerate(rows):
        if _is_empty_row(row):
            skipped += 1
            continue
        if not _label_row(row, AGING_SUMMARY_REPEATED_LABELS):
            errors.append(
                (i + 1, 'سطر خارج بنية ملخص الأعمار (التسميات المتكررة غير مكتملة) — الصف مستبعد', row))
            continue
        rownum = i + 1
        code = normalize_digits(_cell(row, AGING_SUMMARY_VALUE_COLS['customer_code']))
        ccy_raw = _cell(row, AGING_SUMMARY_VALUE_COLS['currency'])
        if not code:
            errors.append((rownum, 'كود عميل ناقص — الصف مستبعد', row))
            continue
        if not ccy_raw:
            errors.append((rownum, 'عملة ناقصة — الصف مستبعد', row))
            continue
        buckets = {}
        for j, bk in enumerate(DETAILS_BUCKET_KEYS):
            num = _bucket_num(_cell(row, AGING_SUMMARY_VALUE_COLS['bucket0'] + j))
            if num is None:
                errors.append(
                    (rownum, f'قيمة غير رقمية في عمود الأعمار ({bk}) — تُعتبر صفرًا', row))
                num = 0.0
            buckets[bk] = num
        records.append({
            'rowNumber': rownum,
            'customerCode': code,
            'customerName': _cell(row, AGING_SUMMARY_VALUE_COLS['customer_name']) or code,
            'currencyRaw': ccy_raw,
            'currency': CURRENCY_MAP.get(ccy_raw, ccy_raw),
            'buckets': buckets,
            'total': to_number(_cell(row, AGING_SUMMARY_VALUE_COLS['total'])),
        })
    return _profile_result(PROFILE_AGING_SUMMARY, rows, 0, errors, records, skipped,
                           'agingSummary')


def parse_debt_aging_block(rows):
    """الأعمار التفصيلية ببنية الكتل: رأس (عميل/عملة/رصيد) + وثائق بفئاتها."""
    records = []
    errors = []
    skipped = 0
    i = 0
    n = len(rows)
    while i < n:
        row = rows[i]
        if _is_empty_row(row):
            skipped += 1
            i += 1
            continue
        if not _is_aging_block_header(row):
            errors.append((i + 1, 'صف خارج كتلة الأعمار التفصيلية — الصف مستبعد', row))
            i += 1
            continue
        rownum = i + 1
        s0 = normalize_text(row[0])
        ccy_raw = s0.split(':', 1)[1].strip() if ':' in s0 else ''
        m = re.match(r'رقم العميل\s*:\s*(\d+)\s*(.*)$', normalize_text(row[3]), re.DOTALL)
        if not m:
            errors.append((rownum, 'رأس كتلة بدون كود عميل — الكتلة مستبعدة', row))
            i += 1
            continue
        code = m.group(1)
        name = normalize_text(m.group(2))
        i += 1
        while i < n and not _is_aging_block_header(rows[i]):
            drow = rows[i]
            if _is_empty_row(drow):
                skipped += 1
                i += 1
                continue
            buckets = {}
            for j, bk in enumerate(DETAILS_BUCKET_KEYS):
                num = _bucket_num(_cell(drow, DETAILS_BUCKET_COLS[j]))
                if num is None:
                    errors.append(
                        (i + 1, f'قيمة غير رقمية في عمود الأعمار ({bk}) — تُعتبر صفرًا', drow))
                    num = 0.0
                buckets[bk] = num
            records.append({
                'rowNumber': i + 1,
                'customerCode': code,
                'customerName': name or code,
                'currencyRaw': ccy_raw,
                'currency': CURRENCY_MAP.get(ccy_raw, ccy_raw),
                'buckets': buckets,
                'total': to_number(_cell(drow, 3)),
                'documentNumber': _cell(drow, 0) or None,
                'documentDate': _cell(drow, 1) or None,
                'documentType': _cell(drow, 2) or None,
            })
            i += 1
    return _profile_result(PROFILE_AGING_DETAILS, rows, 0, errors, records, skipped,
                           'agingDetails')


def parse_profile(rows, profile):
    """توجيه الملف إلى المحلل الصحيح (الكشف مسؤول عنها مسبقًا)."""
    if profile == PROFILE_MASTER:
        for row in rows[:8]:
            if _label_row(row, MASTER_REPEATED_LABELS):
                return parse_customer_master_repeated(rows)
        return parse_customer_master(rows)
    if profile == PROFILE_BALANCE:
        return parse_balance_summary(rows)
    if profile == PROFILE_AGING_SUMMARY:
        for row in rows[:8]:
            if _label_row(row, AGING_SUMMARY_REPEATED_LABELS):
                return parse_aging_summary_repeated(rows)
        return parse_debt_aging(rows, details=False)
    if profile == PROFILE_AGING_DETAILS:
        for row in rows[:8]:
            if _is_aging_block_header(row):
                return parse_debt_aging_block(rows)
        return parse_debt_aging(rows, details=True)
    raise ValueError(f'لا محلل مخصص للملف {profile}')
